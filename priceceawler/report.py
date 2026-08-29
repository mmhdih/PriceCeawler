"""Build daily series from raw price points and export them."""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from typing import Any, Iterable, Sequence

from .jalali import JalaliDate, date_range
from .symbols import Symbol
from .tgju import PricePoint
from .version import APP_NAME, __version__

__all__ = ["Series", "build_series", "to_xlsx", "to_csv", "to_json", "COLUMNS"]

COLUMNS = ("تاریخ شمسی", "روز هفته", "کمترین", "بیشترین", "پایانی", "میانگین معاملاتی", "وضعیت")

_LIVE = "معامله شده"
_CARRIED = "بدون معامله (قیمت روز قبل)"
_MISSING = "بدون داده"


@dataclass
class Series:
    """A symbol's daily rows over a requested Jalali range, plus statistics."""

    symbol: Symbol
    rows: list[dict[str, Any]]
    stats: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {"symbol": self.symbol.to_dict(), "rows": self.rows, "stats": self.stats}


def _round(value: float | None, decimals: int) -> float | int | None:
    """Round to the symbol's precision; whole units come back as ``int``."""
    if value is None:
        return None
    return int(round(value)) if decimals <= 0 else round(value, decimals)


def build_series(
    symbol: Symbol,
    points: Sequence[PricePoint],
    start: JalaliDate,
    end: JalaliDate,
    *,
    fill_gaps: bool = True,
) -> Series:
    """Expand price points into one row per calendar day in ``[start, end]``.

    TGJU has no rows for holidays and weekends. With ``fill_gaps`` the last
    known price is carried forward (flagged in the "وضعیت" column); otherwise
    those days are dropped so the table only holds real trading days.
    """
    decimals = symbol.decimals
    by_date = {point.date: point for point in points}

    # The last observation before the window, so day one is never empty.
    carried: PricePoint | None = None
    start_key = str(start)
    for point in points:
        if point.date < start_key:
            carried = point
        else:
            break

    rows: list[dict[str, Any]] = []
    observed: list[float] = []
    first_close: float | None = None
    last_close: float | None = None

    for day in date_range(start, end):
        key = str(day)
        point = by_date.get(key)
        if point is not None:
            carried = point
            status = _LIVE
        elif fill_gaps and carried is not None:
            point = carried
            status = _CARRIED
        elif fill_gaps:
            point = None
            status = _MISSING
        else:
            continue

        if point is None:
            rows.append(
                {
                    "date": key,
                    "weekday": day.weekday_name,
                    "low": None, "high": None, "close": None, "average": None,
                    "status": status, "live": False,
                }
            )
            continue

        close = _round(point.close, decimals)
        if status == _LIVE and close is not None:
            observed.append(close)
            if first_close is None:
                first_close = close
            last_close = close

        rows.append(
            {
                "date": key,
                "weekday": day.weekday_name,
                "low": _round(point.low, decimals),
                "high": _round(point.high, decimals),
                "close": close,
                "average": _round(point.average, decimals),
                "status": status,
                "live": status == _LIVE,
            }
        )

    change = None
    change_pct = None
    if first_close is not None and last_close is not None and first_close:
        change = _round(last_close - first_close, decimals)
        change_pct = round((last_close - first_close) / first_close * 100, 2)

    stats = {
        "days": len(rows),
        "trading_days": len(observed),
        "first": first_close,
        "last": last_close,
        "min": min(observed) if observed else None,
        "max": max(observed) if observed else None,
        "mean": _round(sum(observed) / len(observed), decimals) if observed else None,
        "change": change,
        "change_pct": change_pct,
        "unit": symbol.unit_label,
    }
    return Series(symbol, rows, stats)


def _row_values(row: dict[str, Any]) -> list[Any]:
    return [
        row["date"], row["weekday"], row["low"], row["high"],
        row["close"], row["average"], row["status"],
    ]


# --------------------------------------------------------------------------
# Exporters
# --------------------------------------------------------------------------

def _sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no ``[]:*?/\\``, unique."""
    cleaned = "".join(" " if c in "[]:*?/\\" else c for c in name).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate, index = cleaned, 2
    while candidate in used:
        suffix = f" ({index})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        index += 1
    used.add(candidate)
    return candidate


def to_xlsx(series_list: Iterable[Series], start: JalaliDate, end: JalaliDate) -> bytes:
    """Render a styled, right-to-left workbook: one sheet per symbol + summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    series_list = list(series_list)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(name="Vazirmatn", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Vazirmatn", size=10)
    muted_font = Font(name="Vazirmatn", size=10, color="8A94A6")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_names: set[str] = set()

    # --- summary sheet ----------------------------------------------------
    summary = workbook.create_sheet(_sheet_name("خلاصه گزارش", used_names))
    summary.sheet_view.rightToLeft = True
    summary["A1"] = f"{APP_NAME} — گزارش قیمت‌های TGJU"
    summary["A1"].font = Font(name="Vazirmatn", bold=True, size=14, color="1F3A5F")
    summary["A2"] = f"بازه: {start} تا {end}   |   نسخه {__version__}"
    summary["A2"].font = muted_font

    summary_header = (
        "نماد", "شناسه TGJU", "واحد", "روزهای معاملاتی",
        "اولین قیمت", "آخرین قیمت", "کمترین", "بیشترین", "میانگین", "تغییر", "درصد تغییر",
    )
    for column, title in enumerate(summary_header, start=1):
        cell = summary.cell(row=4, column=column, value=title)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border

    for offset, series in enumerate(series_list):
        stats = series.stats
        values = (
            series.symbol.name, series.symbol.key, stats["unit"], stats["trading_days"],
            stats["first"], stats["last"], stats["min"], stats["max"], stats["mean"],
            stats["change"], (stats["change_pct"] / 100 if stats["change_pct"] is not None else None),
        )
        for column, value in enumerate(values, start=1):
            cell = summary.cell(row=5 + offset, column=column, value=value)
            cell.font, cell.alignment, cell.border = body_font, center, border
            if column in range(5, 11):
                cell.number_format = "#,##0.####"
            elif column == 11:
                cell.number_format = "0.00%"
    for column in range(1, len(summary_header) + 1):
        summary.column_dimensions[get_column_letter(column)].width = 16 if column > 1 else 24
    summary.freeze_panes = "A5"

    # --- one sheet per symbol --------------------------------------------
    for series in series_list:
        sheet = workbook.create_sheet(_sheet_name(series.symbol.name, used_names))
        sheet.sheet_view.rightToLeft = True
        for column, title in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=1, column=column, value=title)
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border

        number_format = "#,##0" if series.symbol.decimals == 0 else "#,##0." + "0" * series.symbol.decimals
        for index, row in enumerate(series.rows, start=2):
            for column, value in enumerate(_row_values(row), start=1):
                cell = sheet.cell(row=index, column=column, value=value)
                cell.alignment, cell.border = center, border
                cell.font = body_font if row["live"] else muted_font
                if 3 <= column <= 6:
                    cell.number_format = number_format

        widths = (14, 12, 16, 16, 16, 20, 26)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A2"
        if len(series.rows) > 1:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(series.rows) + 1}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_csv(series_list: Iterable[Series]) -> bytes:
    """One flat CSV for every symbol, BOM-prefixed so Excel reads UTF-8."""
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(("نماد", *COLUMNS))
    for series in series_list:
        for row in series.rows:
            writer.writerow((series.symbol.name, *_row_values(row)))
    return "﻿".encode("utf-8") + buffer.getvalue().encode("utf-8")


def to_json(series_list: Iterable[Series], start: JalaliDate, end: JalaliDate) -> bytes:
    payload = {
        "app": APP_NAME,
        "version": __version__,
        "range": {"start": str(start), "end": str(end)},
        "series": [series.to_dict() for series in series_list],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
