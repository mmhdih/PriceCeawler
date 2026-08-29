import io
import json
import unittest

from openpyxl import load_workbook

from priceceawler.jalali import JalaliDate
from priceceawler.report import build_series, to_csv, to_json, to_xlsx
from priceceawler.symbols import CATALOG
from priceceawler.tgju import PricePoint

GOLD = CATALOG["geram18"]


def point(date, close):
    return PricePoint(date, "", None, close - 1000, close + 1000, close)


class TestBuildSeries(unittest.TestCase):
    def setUp(self):
        self.points = [point("1404/05/28", 700_000), point("1404/05/30", 730_000)]
        self.start = JalaliDate(1404, 5, 28)
        self.end = JalaliDate(1404, 6, 1)

    def test_fills_missing_days_with_previous_price(self):
        series = build_series(GOLD, self.points, self.start, self.end)
        self.assertEqual(len(series.rows), 5)
        self.assertEqual(series.rows[1]["close"], 700_000)
        self.assertFalse(series.rows[1]["live"])
        self.assertTrue(series.rows[0]["live"])

    def test_no_fill_keeps_only_trading_days(self):
        series = build_series(GOLD, self.points, self.start, self.end, fill_gaps=False)
        self.assertEqual([row["date"] for row in series.rows], ["1404/05/28", "1404/05/30"])

    def test_stats_use_traded_days_only(self):
        stats = build_series(GOLD, self.points, self.start, self.end).stats
        self.assertEqual(stats["trading_days"], 2)
        self.assertEqual(stats["first"], 700_000)
        self.assertEqual(stats["last"], 730_000)
        self.assertEqual(stats["min"], 700_000)
        self.assertEqual(stats["max"], 730_000)
        self.assertEqual(stats["change"], 30_000)
        self.assertAlmostEqual(stats["change_pct"], 4.29, places=2)
        self.assertEqual(stats["unit"], "تومان")

    def test_price_before_window_is_carried_in(self):
        points = [point("1404/01/01", 500_000), *self.points]
        series = build_series(GOLD, points, JalaliDate(1404, 5, 20), JalaliDate(1404, 5, 28))
        self.assertEqual(series.rows[0]["close"], 500_000)
        self.assertFalse(series.rows[0]["live"])

    def test_days_with_no_prior_data_are_blank(self):
        series = build_series(GOLD, self.points, JalaliDate(1404, 5, 26), self.end)
        self.assertIsNone(series.rows[0]["close"])
        self.assertEqual(series.rows[0]["status"], "بدون داده")

    def test_empty_points_give_empty_stats(self):
        stats = build_series(GOLD, [], self.start, self.end).stats
        self.assertEqual(stats["trading_days"], 0)
        self.assertIsNone(stats["change"])


class TestExporters(unittest.TestCase):
    def setUp(self):
        self.start, self.end = JalaliDate(1404, 5, 28), JalaliDate(1404, 6, 1)
        self.series = [
            build_series(GOLD, [point("1404/05/28", 700_000)], self.start, self.end),
            build_series(CATALOG["ons"], [point("1404/05/28", 3300)], self.start, self.end),
        ]

    def test_xlsx_has_a_summary_and_one_sheet_per_symbol(self):
        workbook = load_workbook(io.BytesIO(to_xlsx(self.series, self.start, self.end)))
        self.assertEqual(workbook.sheetnames[0], "خلاصه گزارش")
        self.assertIn("طلای ۱۸ عیار", workbook.sheetnames)
        sheet = workbook["طلای ۱۸ عیار"]
        self.assertTrue(sheet.sheet_view.rightToLeft)
        self.assertEqual(sheet["A1"].value, "تاریخ شمسی")
        self.assertEqual(sheet["A2"].value, "1404/05/28")
        self.assertEqual(sheet.max_row, 6)  # header + 5 days

    def test_xlsx_sheet_names_stay_within_excel_limits(self):
        long_name = build_series(
            CATALOG["geram18"], [point("1404/05/28", 1)], self.start, self.end
        )
        long_name.symbol = CATALOG["geram18"].__class__(
            "x", "نماد با نام بسیار بسیار بسیار بسیار طولانی برای آزمایش", "گروه"
        )
        workbook = load_workbook(io.BytesIO(to_xlsx([long_name], self.start, self.end)))
        self.assertTrue(all(len(name) <= 31 for name in workbook.sheetnames))

    def test_csv_starts_with_a_bom_for_excel(self):
        data = to_csv(self.series)
        self.assertTrue(data.startswith(b"\xef\xbb\xbf"))
        self.assertIn("طلای ۱۸ عیار", data.decode("utf-8"))

    def test_json_round_trips(self):
        payload = json.loads(to_json(self.series, self.start, self.end))
        self.assertEqual(payload["range"], {"start": "1404/05/28", "end": "1404/06/01"})
        self.assertEqual(len(payload["series"]), 2)
        self.assertEqual(payload["series"][1]["symbol"]["unit"], "دلار")


if __name__ == "__main__":
    unittest.main()
